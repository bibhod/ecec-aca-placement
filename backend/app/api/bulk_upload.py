"""
Bulk Upload API — Issue 17
Upload CSV/Excel for: students, centres, hours, visits/appointments, units.
Also provides GET endpoints to download blank CSV templates.
"""
import csv, io, uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import date
import logging

from app.database import get_db
from app.models import (
    Student, PlacementCentre, HoursLog, Appointment, User,
    QUALIFICATION_CHOICES, UNITS_CHC30125, UNITS_CHC50125,
)
from app.utils.auth import get_current_user

router = APIRouter()
logger = logging.getLogger(__name__)


def _read_file(file_content: bytes, filename: str) -> list:
    """Parse CSV or Excel into a list of dicts."""
    if filename.lower().endswith(".csv"):
        return list(csv.DictReader(io.StringIO(file_content.decode("utf-8-sig"))))
    elif filename.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        return [
            dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
    raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")


def _make_csv_response(headers: list, rows: list, filename: str) -> StreamingResponse:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    out.seek(0)
    return StreamingResponse(
        io.BytesIO(out.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────── TEMPLATES ────────────────────────────

@router.get("/templates/students")
def template_students():
    return _make_csv_response(
        ["student_id", "full_name", "email", "phone", "qualification",
         "campus", "status", "required_hours", "course_start_date", "course_end_date",
         "placement_start_date", "placement_end_date", "notes"],
        [["STU2025001", "Jane Smith", "jane@email.com", "0412345678",
          "CHC30125", "sydney", "active", "160", "2025-02-01", "2025-11-30",
          "2025-04-01", "2025-10-31", ""]],
        "template_students.csv",
    )


@router.get("/templates/centres")
def template_centres():
    return _make_csv_response(
        ["centre_name", "address", "suburb", "state", "postcode",
         "phone", "email", "director_name", "supervisor_name",
         "supervisor_email", "supervisor_phone", "nqs_rating", "max_students"],
        [["Sunshine Childcare", "123 Main St", "Sydney", "NSW", "2000",
          "0291234567", "info@centre.com", "Jane Director",
          "John Trainer", "john@centre.com", "0412345678", "Meeting NQS", "5"]],
        "template_centres.csv",
    )


@router.get("/templates/hours")
def template_hours():
    return _make_csv_response(
        ["student_id", "log_date", "hours", "activity_description"],
        [["STU2025001", "2025-05-01", "8", "Room 2 — toddler activities"]],
        "template_hours.csv",
    )


@router.get("/templates/visits")
def template_visits():
    return _make_csv_response(
        ["student_id", "trainer_assessor_email", "placement_centre_name",
         "visit_date", "visit_time", "duration_hours", "appointment_type",
         "units_assessed", "preparation_notes"],
        [["STU2025001", "trainer@academies.edu.au", "Sunshine Childcare",
          "2025-05-10", "09:00", "1", "cert_iii_1st_visit",
          "Children's Health and Safety; Provide First Aid", ""]],
        "template_visits.csv",
    )


@router.get("/templates/units")
def template_units():
    """Download list of all available units for reference."""
    rows = [["CHC30125", u] for u in UNITS_CHC30125]
    rows += [["CHC50125", u] for u in UNITS_CHC50125]
    return _make_csv_response(
        ["qualification", "unit_name"], rows, "template_units_reference.csv",
    )


# ─────────────────────────────────────── IMPORT ───────────────────────────────

@router.post("/import/students")
async def import_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    rows = _read_file(content, file.filename)
    created, skipped, errors = [], [], []
    for i, row in enumerate(rows, 2):
        sid = (row.get("student_id") or "").strip()
        name = (row.get("full_name") or "").strip()
        qual = (row.get("qualification") or "").strip()
        campus = (row.get("campus") or "sydney").strip()
        if not sid or not name or not qual:
            errors.append({"row": i, "error": "Missing student_id, full_name or qualification"}); continue
        if qual not in QUALIFICATION_CHOICES:
            errors.append({"row": i, "error": f"Invalid qualification '{qual}'"}); continue
        if db.query(Student).filter(Student.student_id == sid).first():
            skipped.append(sid); continue
        try:
            hrs = float(row.get("required_hours") or 0) or (288 if "50" in qual else 160)
            s = Student(
                student_id=sid, full_name=name,
                email=row.get("email") or None, phone=row.get("phone") or None,
                qualification=qual, campus=campus,
                status=row.get("status") or "active",
                required_hours=hrs, completed_hours=0,
                course_start_date=date.fromisoformat(row["course_start_date"]) if row.get("course_start_date") else None,
                course_end_date=date.fromisoformat(row["course_end_date"]) if row.get("course_end_date") else None,
                placement_start_date=date.fromisoformat(row["placement_start_date"]) if row.get("placement_start_date") else None,
                placement_end_date=date.fromisoformat(row["placement_end_date"]) if row.get("placement_end_date") else None,
                notes=row.get("notes") or None,
            )
            db.add(s); created.append(sid)
        except Exception as e:
            errors.append({"row": i, "student_id": sid, "error": str(e)})
    db.commit()
    return {"message": f"{len(created)} created, {len(skipped)} skipped, {len(errors)} errors",
            "created": created, "skipped": skipped, "errors": errors}


@router.post("/import/centres")
async def import_centres(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    rows = _read_file(content, file.filename)
    created, errors = [], []
    for i, row in enumerate(rows, 2):
        name = (row.get("centre_name") or "").strip()
        if not name:
            errors.append({"row": i, "error": "centre_name required"}); continue
        try:
            c = PlacementCentre(
                centre_name=name,
                address=row.get("address") or None, suburb=row.get("suburb") or None,
                state=row.get("state") or None, postcode=row.get("postcode") or None,
                phone=row.get("phone") or None, email=row.get("email") or None,
                director_name=row.get("director_name") or None,
                supervisor_name=row.get("supervisor_name") or None,
                supervisor_email=row.get("supervisor_email") or None,
                supervisor_phone=row.get("supervisor_phone") or None,
                nqs_rating=row.get("nqs_rating") or None,
                max_students=int(row.get("max_students") or 5),
                approved=True,
            )
            db.add(c); created.append(name)
        except Exception as e:
            errors.append({"row": i, "name": name, "error": str(e)})
    db.commit()
    return {"message": f"{len(created)} centres created, {len(errors)} errors",
            "created": created, "errors": errors}


@router.post("/import/hours")
async def import_hours(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    rows = _read_file(content, file.filename)
    created, errors = [], []
    for i, row in enumerate(rows, 2):
        sid = (row.get("student_id") or "").strip()
        log_date = (row.get("log_date") or "").strip()
        hrs = (row.get("hours") or "").strip()
        if not sid or not log_date or not hrs:
            errors.append({"row": i, "error": "Missing required fields"}); continue
        student = db.query(Student).filter(Student.student_id == sid).first()
        if not student:
            errors.append({"row": i, "error": f"Student '{sid}' not found"}); continue
        try:
            h = float(hrs)
            log = HoursLog(
                student_id=student.id, log_date=date.fromisoformat(log_date),
                hours=h, activity_description=row.get("activity_description") or None,
                flagged_unrealistic=h > 10, approved=False, created_by=current_user.id,
            )
            db.add(log)
            student.completed_hours = (student.completed_hours or 0) + h
            created.append(f"{sid}/{log_date}")
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    db.commit()
    return {"message": f"{len(created)} entries created, {len(errors)} errors",
            "created": created, "errors": errors}


@router.post("/import/visits")
async def import_visits(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    rows = _read_file(content, file.filename)
    created, errors = [], []
    for i, row in enumerate(rows, 2):
        sid = (row.get("student_id") or "").strip()
        vdate = (row.get("visit_date") or "").strip()
        if not sid or not vdate:
            errors.append({"row": i, "error": "Missing student_id or visit_date"}); continue
        student = db.query(Student).filter(Student.student_id == sid).first()
        if not student:
            errors.append({"row": i, "error": f"Student '{sid}' not found"}); continue
        trainer = None
        if row.get("trainer_assessor_email"):
            trainer = db.query(User).filter(User.email == row["trainer_assessor_email"].strip()).first()
        centre = None
        if row.get("placement_centre_name"):
            centre = db.query(PlacementCentre).filter(
                PlacementCentre.centre_name.ilike(f"%{row['placement_centre_name'].strip()}%")
            ).first()
        units = [u.strip() for u in row.get("units_assessed", "").split(";") if u.strip()]
        try:
            a = Appointment(
                student_id=student.id,
                trainer_assessor_id=trainer.id if trainer else current_user.id,
                title=f"{row.get('appointment_type','visit')} – {student.full_name}",
                appointment_type=row.get("appointment_type") or "cert_iii_1st_visit",
                placement_centre_id=centre.id if centre else None,
                scheduled_date=date.fromisoformat(vdate),
                scheduled_time=row.get("visit_time") or "09:00",
                duration_hours=float(row.get("duration_hours") or 1),
                units_assessed=units,
                preparation_notes=row.get("preparation_notes") or None,
                status="scheduled",
                visit_reference=f"VIS-{uuid.uuid4().hex[:8].upper()}",
                created_by=current_user.id,
            )
            db.add(a); created.append(f"{sid}/{vdate}")
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    db.commit()
    return {"message": f"{len(created)} visits created, {len(errors)} errors",
            "created": created, "errors": errors}
