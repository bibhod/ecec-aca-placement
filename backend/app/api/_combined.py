"""
Combined routers for centres, notifications, and reports.
Fixes:
  Issue 11 — NQS label change (API returns field; label is in the UI)
  Issue 12 — export functions produce valid CSV output
  Issue 17 — bulk import for centres
"""
import csv, io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional
from pydantic import BaseModel
from datetime import date

from app.database import get_db
from app.models import PlacementCentre, Student, Notification, User, HoursLog, ComplianceDocument, Appointment
from app.utils.auth import get_current_user

# ─── CENTRES ────────────────────────────────────────────────────────────────
centres_router = APIRouter()


def centre_to_dict(c: PlacementCentre) -> dict:
    return {
        "id": c.id,
        "centre_name": c.centre_name,
        "address": c.address,
        "suburb": c.suburb,
        "state": c.state,
        "postcode": c.postcode,
        "latitude": c.latitude,
        "longitude": c.longitude,
        "phone": c.phone,
        "email": c.email,
        "director_name": c.director_name,
        "director_email": c.director_email,
        "supervisor_name": c.supervisor_name,
        "supervisor_email": c.supervisor_email,
        "supervisor_phone": c.supervisor_phone,
        # Issue 11 — field name unchanged; UI renames the label
        "nqs_rating": c.nqs_rating,
        "max_students": c.max_students,
        "accepted_qualifications": c.accepted_qualifications,
        "approved": c.approved,
        "notes": c.notes,
        "created_at": str(c.created_at) if c.created_at else None,
    }


@centres_router.get("")
def list_centres(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    centres = db.query(PlacementCentre).order_by(PlacementCentre.centre_name).all()
    result = []
    for c in centres:
        d = centre_to_dict(c)
        d["student_count"] = db.query(Student).filter(
            Student.placement_centre_id == c.id, Student.status == "current"
        ).count()
        result.append(d)
    return result


@centres_router.get("/{centre_id}")
def get_centre(centre_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(PlacementCentre).filter(PlacementCentre.id == centre_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Centre not found")
    return centre_to_dict(c)


class CentreCreate(BaseModel):
    centre_name: str
    address: Optional[str] = None
    suburb: Optional[str] = None
    state: Optional[str] = None
    postcode: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    director_name: Optional[str] = None
    director_email: Optional[str] = None
    supervisor_name: Optional[str] = None
    supervisor_email: Optional[str] = None
    supervisor_phone: Optional[str] = None
    nqs_rating: Optional[str] = None
    max_students: int = 5
    accepted_qualifications: Optional[list] = None
    approved: bool = True
    notes: Optional[str] = None


@centres_router.post("")
def create_centre(data: CentreCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = PlacementCentre(**data.dict())
    db.add(c)
    db.commit()
    db.refresh(c)
    return centre_to_dict(c)


@centres_router.put("/{centre_id}")
def update_centre(centre_id: str, data: CentreCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(PlacementCentre).filter(PlacementCentre.id == centre_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Centre not found")
    for k, v in data.dict(exclude_none=True).items():
        if hasattr(c, k):
            setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return centre_to_dict(c)


@centres_router.delete("/{centre_id}")
def delete_centre(centre_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(PlacementCentre).filter(PlacementCentre.id == centre_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Centre not found")
    db.delete(c)
    db.commit()
    return {"message": "Centre deleted"}


@centres_router.post("/bulk-import")
async def bulk_import_centres(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk-import placement centres from CSV (Issue 17).
    Expected CSV headers: centre_name, address, suburb, state, postcode,
    phone, email, director_name, supervisor_name, supervisor_email,
    supervisor_phone, nqs_rating
    """
    content = await file.read()
    if file.filename.lower().endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        rows = list(reader)
    elif file.filename.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])) for row in ws.iter_rows(min_row=2, values_only=True)]
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    created, errors = [], []
    for i, row in enumerate(rows, start=2):
        name = row.get("centre_name", "").strip()
        if not name:
            errors.append({"row": i, "error": "centre_name is required"})
            continue
        try:
            c = PlacementCentre(
                centre_name=name,
                address=row.get("address") or None,
                suburb=row.get("suburb") or None,
                state=row.get("state") or None,
                postcode=row.get("postcode") or None,
                phone=row.get("phone") or None,
                email=row.get("email") or None,
                director_name=row.get("director_name") or None,
                supervisor_name=row.get("supervisor_name") or None,
                supervisor_email=row.get("supervisor_email") or None,
                supervisor_phone=row.get("supervisor_phone") or None,
                nqs_rating=row.get("nqs_rating") or None,
                approved=True,
            )
            db.add(c)
            created.append(name)
        except Exception as e:
            errors.append({"row": i, "name": name, "error": str(e)})

    db.commit()
    return {
        "message": f"Import complete: {len(created)} created, {len(errors)} errors",
        "created": created,
        "errors": errors,
    }


# ─── NOTIFICATIONS ──────────────────────────────────────────────────────────
notifications_router = APIRouter()


@notifications_router.get("")
def list_notifications(
    unread_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Notification).filter(Notification.user_id == current_user.id)
    if unread_only:
        q = q.filter(Notification.read == False)
    notes = q.order_by(Notification.created_at.desc()).limit(50).all()
    return [
        {
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "type": n.type,
            "link": n.link,
            "read": n.read,
            "created_at": str(n.created_at) if n.created_at else None,
        }
        for n in notes
    ]


@notifications_router.put("/read-all")
def mark_all_read(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    db.query(Notification).filter(
        Notification.user_id == current_user.id, Notification.read == False
    ).update({"read": True})
    db.commit()
    return {"message": "All notifications marked as read"}


@notifications_router.put("/{notif_id}/read")
def mark_read(notif_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    n = db.query(Notification).filter(
        Notification.id == notif_id, Notification.user_id == current_user.id
    ).first()
    if not n:
        raise HTTPException(status_code=404, detail="Notification not found")
    n.read = True
    db.commit()
    return {"message": "Marked as read"}


# ─── REPORTS ────────────────────────────────────────────────────────────────
reports_router = APIRouter()

# Qualification labels (Issue 9)
QUAL_LABELS = {
    "CHC30121": "Cert III (Superseded)",
    "CHC50121": "Diploma (Superseded)",
    "CHC30125": "Cert III",
    "CHC50125": "Diploma",
}


@reports_router.get("/overview")
def report_overview(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    today = date.today()
    students = db.query(Student).all()

    by_campus, by_qualification, by_status = {}, {}, {}
    hours_data, compliance_data = [], []

    for s in students:
        by_campus[s.campus] = by_campus.get(s.campus, 0) + 1
        qual_label = QUAL_LABELS.get(s.qualification, s.qualification)
        by_qualification[qual_label] = by_qualification.get(qual_label, 0) + 1
        by_status[s.status] = by_status.get(s.status, 0) + 1

        pct = round(s.completed_hours / s.required_hours * 100, 1) if s.required_hours else 0
        hours_data.append({
            "student_id": s.student_id,
            "student_name": s.full_name,
            "campus": s.campus,
            "qualification": QUAL_LABELS.get(s.qualification, s.qualification),
            "required_hours": s.required_hours,
            "completed_hours": s.completed_hours,
            "percentage": pct,
            "status": s.status,
        })

        docs = db.query(ComplianceDocument).filter(ComplianceDocument.student_id == s.id).all()
        expired = sum(1 for d in docs if d.expiry_date and d.expiry_date < today)
        expiring = sum(1 for d in docs if d.expiry_date and today <= d.expiry_date <= date.fromordinal(today.toordinal() + 30))
        pending = sum(1 for d in docs if not d.verified)
        compliance_data.append({
            "student_id": s.student_id,
            "student_name": s.full_name,
            "campus": s.campus,
            "total_docs": len(docs),
            "expired": expired,
            "expiring_soon": expiring,
            "pending_verification": pending,
            "compliant": len(docs) > 0 and expired == 0 and pending == 0,
        })

    return {
        "by_campus": by_campus,
        "by_qualification": by_qualification,
        "by_status": by_status,
        "hours_data": sorted(hours_data, key=lambda x: x["percentage"], reverse=True),
        "compliance_data": compliance_data,
        "summary": {
            "total_students": len(students),
            "total_hours_completed": sum(s.completed_hours for s in students),
            "total_hours_required": sum(s.required_hours for s in students),
            "compliance_rate": round(
                sum(1 for c in compliance_data if c["compliant"]) / len(compliance_data) * 100, 1
            ) if compliance_data else 0,
        },
    }


@reports_router.get("/export/students")
def export_students_csv(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export all students to CSV (Issue 12)."""
    from fastapi.responses import StreamingResponse
    students = db.query(Student).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Student ID", "Full Name", "Email", "Phone", "Qualification", "Campus",
        "Status", "Required Hours", "Completed Hours", "Progress %",
        "Placement Centre", "Course Start", "Course End",
    ])
    for s in students:
        centre = s.placement_centre
        pct = round(s.completed_hours / s.required_hours * 100, 1) if s.required_hours else 0
        writer.writerow([
            s.student_id, s.full_name, s.email or "", s.phone or "",
            s.qualification, s.campus, s.status,
            s.required_hours, s.completed_hours, f"{pct}%",
            centre.centre_name if centre else "",
            str(s.course_start_date) if s.course_start_date else "",
            str(s.course_end_date) if s.course_end_date else "",
        ])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_report.csv"},
    )


@reports_router.get("/export/hours")
def export_hours_csv(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export all hours logs to CSV (Issue 12)."""
    from fastapi.responses import StreamingResponse
    logs = db.query(HoursLog).order_by(HoursLog.log_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Student ID", "Student Name", "Hours", "Activity", "Approved", "Approved By", "Flagged"])
    for l in logs:
        student = db.query(Student).filter(Student.id == l.student_id).first()
        flags = []
        if getattr(l, "flagged_unrealistic", False):
            flags.append("unrealistic")
        if getattr(l, "flagged_duplicate", False):
            flags.append("duplicate")
        writer.writerow([
            str(l.log_date), student.student_id if student else "", student.full_name if student else "",
            l.hours, l.activity_description or "", "Yes" if l.approved else "No",
            l.approved_by or "", ",".join(flags) if flags else "",
        ])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=hours_report.csv"},
    )


@reports_router.get("/export/audit")
def export_audit_csv(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Export audit log to CSV (Issue 14)."""
    from fastapi.responses import StreamingResponse
    from app.models import AuditLog
    entries = db.query(AuditLog).order_by(AuditLog.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "User Email", "User Name", "Action", "Resource Type", "Resource ID", "Resource Label"])
    for e in entries:
        writer.writerow([
            str(e.created_at), e.user_email or "", e.user_name or "",
            e.action, e.resource_type or "", e.resource_id or "", e.resource_label or "",
        ])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=audit_report.csv"},
    )


@reports_router.get("/export/pdf")
def export_report_pdf(
    report_type: str = "enrollment_summary",
    campus: str = "",
    qualification: str = "",
    status: str = "current",
    days: str = "30",
    missing_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generate a PDF export of any Custom Report.
    Uses the Academies Australasia brand header from the email templates.
    Returns file as attachment download.
    """
    from fastapi.responses import Response
    from datetime import datetime as dt
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io as _io
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="ReportLab not installed. Run: pip install reportlab")

    today = date.today()
    now_str = dt.now().strftime("%d %b %Y %H:%M")

    # ── Brand colours ────────────────────────────────────────────────────────
    NAVY   = colors.HexColor("#1A2B5F")
    CYAN   = colors.HexColor("#00AEEF")
    LGRAY  = colors.HexColor("#F0F4F8")
    DGRAY  = colors.HexColor("#2C3E50")
    WHITE  = colors.white

    # ── Fetch data ───────────────────────────────────────────────────────────
    rows_data, headers, title, filter_desc = [], [], "", ""

    def _qual_match(s_qual):
        if not qualification:
            return True
        if qualification == "cert_iii":
            return "30" in (s_qual or "")
        if qualification == "diploma":
            return "50" in (s_qual or "")
        return s_qual == qualification

    if report_type == "enrollment_summary":
        title = "Student Enrollment Overview"
        filter_parts = []
        if campus:   filter_parts.append(f"Campus: {campus.title()}")
        if qualification: filter_parts.append(f"Qualification: {qualification}")
        if status:   filter_parts.append(f"Status: {status}")
        filter_desc = "  |  ".join(filter_parts) if filter_parts else "All students"

        q = db.query(Student)
        if status:
            q = q.filter(Student.status == status)
        students = q.all()
        if campus:
            students = [s for s in students if (s.campus or "").lower() == campus.lower()]
        if qualification:
            students = [s for s in students if _qual_match(s.qualification)]

        headers = ["Student ID", "Name", "Campus", "Qualification", "Status", "Compliance", "Hours %"]
        for s in students:
            pct = round(s.completed_hours / s.required_hours * 100, 0) if s.required_hours else 0
            rows_data.append([
                s.student_id, s.full_name, (s.campus or "").title(),
                s.qualification or "—", s.status.title(),
                "—",  # compliance computed separately if needed
                f"{pct:.0f}%",
            ])

    elif report_type == "placement_hours":
        title = "Placement Hours Summary"
        filter_parts = []
        if campus:   filter_parts.append(f"Campus: {campus.title()}")
        if qualification: filter_parts.append(f"Qualification: {qualification}")
        filter_desc = "  |  ".join(filter_parts) if filter_parts else "All active students"

        q = db.query(Student)
        if status:
            q = q.filter(Student.status == status)
        students = q.all()
        if campus:
            students = [s for s in students if (s.campus or "").lower() == campus.lower()]
        if qualification:
            students = [s for s in students if _qual_match(s.qualification)]
        students = sorted(students, key=lambda s: s.completed_hours / (s.required_hours or 1))

        headers = ["Student ID", "Name", "Campus", "Qualification", "Completed", "Required", "Progress"]
        for s in students:
            pct = round(s.completed_hours / s.required_hours * 100, 0) if s.required_hours else 0
            rows_data.append([
                s.student_id, s.full_name, (s.campus or "").title(),
                s.qualification or "—",
                f"{s.completed_hours:.0f}h", f"{s.required_hours:.0f}h",
                f"{pct:.0f}%",
            ])

    elif report_type == "expiring_documents":
        title = "Expiring Documents Report"
        days_int = int(days) if days.isdigit() else 30
        filter_desc = f"Documents expiring within {days_int} days"
        expiry_limit = today + __import__("datetime").timedelta(days=days_int)
        docs = db.query(ComplianceDocument).filter(
            ComplianceDocument.expiry_date >= today,
            ComplianceDocument.expiry_date <= expiry_limit,
        ).order_by(ComplianceDocument.expiry_date).all()

        headers = ["Student", "Campus", "Document Type", "Expiry Date", "Days Left", "Verified"]
        for d in docs:
            s = db.query(Student).filter(Student.id == d.student_id).first()
            days_left = (d.expiry_date - today).days
            doc_label = d.document_type.replace("_", " ").title()
            rows_data.append([
                s.full_name if s else "Unknown",
                (s.campus or "—").title() if s else "—",
                doc_label,
                str(d.expiry_date),
                f"{days_left} days",
                "Yes" if d.verified else "No",
            ])

    elif report_type == "compliance_status":
        title = "Compliance Status Report"
        filter_parts = []
        if campus:       filter_parts.append(f"Campus: {campus.title()}")
        if missing_only: filter_parts.append("Incomplete only")
        filter_desc = "  |  ".join(filter_parts) if filter_parts else "All active students"

        students = db.query(Student).filter(Student.status == "current").all()
        if campus:
            students = [s for s in students if (s.campus or "").lower() == campus.lower()]

        REQUIRED_4 = ["working_with_children_check", "first_aid_certificate",
                      "work_placement_agreement", "memorandum_of_understanding"]
        headers = ["Student ID", "Name", "Campus", "Qual", "Docs Submitted", "Status", "Outstanding"]
        for s in students:
            docs = db.query(ComplianceDocument).filter(ComplianceDocument.student_id == s.id).all()
            submitted = {d.document_type for d in docs}
            missing = [t.replace("_", " ").upper()[:4] for t in REQUIRED_4 if t not in submitted]
            if missing_only and not missing:
                continue
            rows_data.append([
                s.student_id, s.full_name, (s.campus or "").title(), s.qualification or "—",
                f"{len(REQUIRED_4)-len(missing)}/{len(REQUIRED_4)}",
                "Complete" if not missing else "Incomplete",
                ", ".join(missing) if missing else "✓ All submitted",
            ])

    else:
        title = "Custom Report"
        filter_desc = ""
        headers = ["No data"]
        rows_data = [["Unknown report type"]]

    # ── Build PDF ─────────────────────────────────────────────────────────────
    buffer = _io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    # --- Header block ---
    header_style = ParagraphStyle(
        "header", parent=styles["Normal"],
        fontSize=18, textColor=WHITE, fontName="Helvetica-Bold", spaceAfter=4,
    )
    tagline_style = ParagraphStyle(
        "tagline", parent=styles["Normal"],
        fontSize=9, textColor=CYAN, fontName="Helvetica", spaceAfter=0,
    )
    header_table = Table(
        [[
            Paragraph("Academies Australasia", header_style),
            "",
        ], [
            Paragraph("ECEC Work Placement Management System", tagline_style),
            "",
        ]],
        colWidths=["70%", "30%"],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), NAVY),
        ("PADDING",     (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.4*cm))

    # --- Report title + metadata ---
    title_style = ParagraphStyle(
        "rptTitle", parent=styles["Normal"],
        fontSize=14, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "meta", parent=styles["Normal"],
        fontSize=9, textColor=DGRAY, spaceAfter=2,
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated: {now_str}  |  Rows: {len(rows_data)}", meta_style))
    if filter_desc:
        elements.append(Paragraph(f"Filters: {filter_desc}", meta_style))
    elements.append(Spacer(1, 0.4*cm))

    # --- Data table ---
    if rows_data:
        col_count = len(headers)
        page_w = A4[0] - 3*cm
        col_w = page_w / col_count

        table_data = [headers] + rows_data
        tbl = Table(table_data, colWidths=[col_w] * col_count, repeatRows=1)
        tbl.setStyle(TableStyle([
            # Header row
            ("BACKGROUND",   (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR",    (0, 0), (-1, 0), WHITE),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 8),
            ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING",(0, 0), (-1, 0), 8),
            ("TOPPADDING",   (0, 0), (-1, 0), 8),
            # Data rows
            ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LGRAY]),
            ("ALIGN",        (0, 1), (-1, -1), "LEFT"),
            ("TOPPADDING",   (0, 1), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 1), (-1, -1), 5),
            # Grid
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#DCE6EF")),
            ("LINEBELOW",    (0, 0), (-1, 0), 1.5, CYAN),
        ]))
        elements.append(tbl)
    else:
        elements.append(Paragraph("No data found for the selected filters.", meta_style))

    # --- Footer ---
    elements.append(Spacer(1, 0.5*cm))
    footer_style = ParagraphStyle(
        "footer", parent=styles["Normal"],
        fontSize=7, textColor=colors.HexColor("#9AAAB8"),
    )
    elements.append(Paragraph(
        "Academies Australasia  |  Level 6, 505 George Street, Sydney NSW 2000  |  "
        "T: +61 2 9224 5500  |  www.academies.edu.au  |  "
        "© 2026 Academies Australasia. All rights reserved.",
        footer_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    filename = f"{report_type}_{today}.pdf"
    return Response(
        content=buffer.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
