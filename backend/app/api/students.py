"""
Students API
Fixes:
  Issue 9  — support all four qualifications (CHC30121/50121 superseded + CHC30125/50125)
  Issue 13 — bulk import from CSV/Excel
  Issue 17 — bulk upload endpoint
"""
import csv, io, uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional, List
from pydantic import BaseModel
from datetime import date

from app.database import get_db
from app.models import Student, PlacementCentre, ComplianceDocument, HoursLog, User, QUALIFICATION_CHOICES
from app.utils.auth import get_current_user
from app.api.audit import write_audit

router = APIRouter()


def student_to_dict(s: Student, db: Session) -> dict:
    centre = s.placement_centre
    docs = db.query(ComplianceDocument).filter(ComplianceDocument.student_id == s.id).all()
    today = date.today()

    # A student is only compliant when ALL 4 required doc types are submitted
    REQUIRED_4 = ['working_with_children_check', 'first_aid_certificate',
                  'work_placement_agreement', 'memorandum_of_understanding']
    submitted_types = {d.document_type for d in docs}
    required_submitted_count = sum(1 for t in REQUIRED_4 if t in submitted_types)
    missing_count = len(REQUIRED_4) - required_submitted_count

    if missing_count > 0:
        compliance_status = "pending"
    else:
        # All 4 submitted — check expiry on the latest per required type
        latest_docs: dict = {}
        for d in docs:
            if d.document_type in REQUIRED_4:
                existing = latest_docs.get(d.document_type)
                if not existing or (d.created_at or date.min) > (existing.created_at or date.min):
                    latest_docs[d.document_type] = d
        expired = any(d.expiry_date and d.expiry_date < today for d in latest_docs.values())
        compliance_status = "expired" if expired else "compliant"

    return {
        "id": s.id,
        "student_id": s.student_id,
        "full_name": s.full_name,
        "email": s.email,
        "phone": s.phone,
        "date_of_birth": str(s.date_of_birth) if s.date_of_birth else None,
        "qualification": s.qualification,
        "campus": s.campus,
        "status": s.status,
        "course_start_date": str(s.course_start_date) if s.course_start_date else None,
        "course_end_date": str(s.course_end_date) if s.course_end_date else None,
        "placement_centre_id": s.placement_centre_id,
        "placement_start_date": str(s.placement_start_date) if s.placement_start_date else None,
        "placement_end_date": str(s.placement_end_date) if s.placement_end_date else None,
        "required_hours": s.required_hours,
        "completed_hours": s.completed_hours,
        "hours_percentage": round((s.completed_hours / s.required_hours * 100) if s.required_hours else 0, 1),
        "coordinator_id": s.coordinator_id,
        "preferred_suburb": getattr(s, "preferred_suburb", None),
        "preferred_state": getattr(s, "preferred_state", None),
        "notes": s.notes,
        "compliance_status": compliance_status,
        "compliance_submitted_count": required_submitted_count,
        "compliance_missing_count": missing_count,
        "placement_site": {
            "id": centre.id,
            "centre_name": centre.centre_name,
            "address": ", ".join(filter(None, [centre.address, centre.suburb, centre.state, centre.postcode])),
            "supervisor_name": centre.supervisor_name,
            "supervisor_email": centre.supervisor_email,
            "supervisor_phone": centre.supervisor_phone,
        } if centre else None,
        "compliance_documents": [
            {
                "id": d.id,
                "document_type": d.document_type,
                "document_number": d.document_number,
                "issue_date": str(d.issue_date) if d.issue_date else None,
                "expiry_date": str(d.expiry_date) if d.expiry_date else None,
                "verified": d.verified,
                "verified_by": d.verified_by,
                "verified_at": str(d.verified_at) if d.verified_at else None,
                "file_url": d.file_url,
                "notes": d.notes,
                "status": (
                    "expired" if d.expiry_date and d.expiry_date < today
                    else "expiring_soon" if d.expiry_date and (d.expiry_date - today).days <= 30
                    else "valid" if d.verified
                    else "pending"
                ),
            }
            for d in docs
        ],
        "created_at": str(s.created_at) if s.created_at else None,
    }


@router.get("")
def list_students(
    search: Optional[str] = None,
    campus: Optional[str] = None,
    qualification: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(Student)
    if search:
        q = q.filter(or_(
            Student.full_name.ilike(f"%{search}%"),
            Student.student_id.ilike(f"%{search}%"),
            Student.email.ilike(f"%{search}%"),
        ))
    if campus:
        q = q.filter(Student.campus == campus)
    if qualification:
        q = q.filter(Student.qualification == qualification)
    if status:
        q = q.filter(Student.status == status)
    students = q.order_by(Student.full_name).all()
    return [student_to_dict(s, db) for s in students]


@router.get("/qualifications")
def get_qualifications():
    """Return all valid qualification codes and labels (Issue 9)."""
    labels = {
        "CHC30121": "CHC30121 – Certificate III in Early Childhood Education and Care (Superseded)",
        "CHC50121": "CHC50121 – Diploma of Early Childhood Education and Care (Superseded)",
        "CHC30125": "CHC30125 – Certificate III in Early Childhood Education and Care",
        "CHC50125": "CHC50125 – Diploma of Early Childhood Education and Care",
    }
    return [{"value": k, "label": v} for k, v in labels.items()]


@router.get("/{student_id}")
def get_student(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    return student_to_dict(s, db)


class StudentCreate(BaseModel):
    student_id: str
    full_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    qualification: str
    campus: str
    status: str = "active"
    course_start_date: Optional[str] = None
    course_end_date: Optional[str] = None
    placement_centre_id: Optional[str] = None
    placement_start_date: Optional[str] = None
    placement_end_date: Optional[str] = None
    required_hours: float = 160
    coordinator_id: Optional[str] = None
    preferred_suburb: Optional[str] = None
    preferred_state: Optional[str] = None
    notes: Optional[str] = None


@router.post("")
def create_student(
    data: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing = db.query(Student).filter(Student.student_id == data.student_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student ID already exists")

    # Validate qualification (Issue 9)
    if data.qualification not in QUALIFICATION_CHOICES:
        raise HTTPException(status_code=400, detail=f"Invalid qualification. Valid: {QUALIFICATION_CHOICES}")

    # Auto-set hours based on qualification
    required_hours = data.required_hours
    if data.qualification in ("CHC50121", "CHC50125") and required_hours == 160:
        required_hours = 288

    s = Student(
        student_id=data.student_id,
        full_name=data.full_name,
        email=data.email,
        phone=data.phone,
        date_of_birth=date.fromisoformat(data.date_of_birth) if data.date_of_birth else None,
        qualification=data.qualification,
        campus=data.campus,
        status=data.status,
        course_start_date=date.fromisoformat(data.course_start_date) if data.course_start_date else None,
        course_end_date=date.fromisoformat(data.course_end_date) if data.course_end_date else None,
        placement_centre_id=data.placement_centre_id,
        placement_start_date=date.fromisoformat(data.placement_start_date) if data.placement_start_date else None,
        placement_end_date=date.fromisoformat(data.placement_end_date) if data.placement_end_date else None,
        required_hours=required_hours,
        completed_hours=0,
        coordinator_id=data.coordinator_id,
        preferred_suburb=data.preferred_suburb,
        preferred_state=data.preferred_state,
        notes=data.notes,
    )
    db.add(s)
    db.commit()
    db.refresh(s)

    if s.email:
        from app.services.email_service import email_welcome_student
        coordinator = db.query(User).filter(User.id == s.coordinator_id).first() if s.coordinator_id else current_user
        email_welcome_student(
            student_name=s.full_name,
            student_email=s.email,
            student_id=s.student_id,
            qualification=s.qualification,
            campus=s.campus,
            coordinator_name=coordinator.full_name if coordinator else current_user.full_name,
            coordinator_email=coordinator.email if coordinator else current_user.email,
            frontend_url="",
        )

    # Audit: record student creation
    write_audit(
        db, current_user, "student.create", "student",
        resource_id=s.id, resource_label=f"{s.full_name} ({s.student_id})",
        details={"student_id": s.student_id, "qualification": s.qualification, "campus": s.campus},
    )
    db.commit()

    return student_to_dict(s, db)


@router.put("/{student_id}")
def update_student(
    student_id: str,
    data: StudentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")

    if data.qualification not in QUALIFICATION_CHOICES:
        raise HTTPException(status_code=400, detail=f"Invalid qualification. Valid: {QUALIFICATION_CHOICES}")

    date_fields = {
        "course_start_date", "course_end_date",
        "placement_start_date", "placement_end_date", "date_of_birth",
    }
    for field, val in data.dict(exclude_none=True).items():
        if field in date_fields:
            setattr(s, field, date.fromisoformat(val) if val else None)
        else:
            if hasattr(s, field):
                setattr(s, field, val)

    db.commit()
    db.refresh(s)

    # Audit: record student update
    write_audit(
        db, current_user, "student.update", "student",
        resource_id=s.id, resource_label=f"{s.full_name} ({s.student_id})",
        details={"updated_fields": list(data.dict(exclude_none=True).keys())},
    )
    db.commit()

    return student_to_dict(s, db)


# ─── Placement Completion Checklist ──────────────────────────────────────────

@router.get("/{student_id}/checklist")
def get_placement_checklist(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns the real-time placement completion checklist for a student.
    All four criteria must be green before a completion record can be generated.
    """
    from app.models import ComplianceDocument, HoursLog, Appointment, Issue, PlacementCompletion, VISIT_LIMITS
    from datetime import date

    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")

    today = date.today()
    REQUIRED_4 = [
        "working_with_children_check", "first_aid_certificate",
        "work_placement_agreement", "memorandum_of_understanding",
    ]

    # 1. All 4 compliance docs submitted (not just any doc — all required types)
    submitted_types = {
        d.document_type
        for d in db.query(ComplianceDocument).filter(
            ComplianceDocument.student_id == s.id
        ).all()
    }
    compliance_ok = all(t in submitted_types for t in REQUIRED_4)
    compliance_detail = (
        "All 4 required documents submitted"
        if compliance_ok
        else f"Missing: {', '.join(t.replace('_', ' ').title() for t in REQUIRED_4 if t not in submitted_types)}"
    )

    # 2. Required placement hours met
    hours_ok = s.completed_hours >= s.required_hours if s.required_hours else False
    hours_detail = (
        f"{s.completed_hours:.0f} / {s.required_hours:.0f} hours completed"
    )

    # 3. All required visits completed (status = completed or approved)
    required_visits = VISIT_LIMITS.get(s.qualification, 3)
    completed_visits = db.query(Appointment).filter(
        Appointment.student_id == s.id,
        Appointment.completed == True,
    ).count()
    visits_ok = completed_visits >= required_visits
    visits_detail = f"{completed_visits} of {required_visits} required visits completed"

    # 4. No open critical issues
    open_critical = db.query(Issue).filter(
        Issue.student_id == s.id,
        Issue.status.in_(["open", "in_progress"]),
        Issue.priority == "critical",
    ).count()
    issues_ok = open_critical == 0
    issues_detail = (
        "No open critical issues"
        if issues_ok
        else f"{open_critical} open critical issue(s) must be resolved"
    )

    all_green = compliance_ok and hours_ok and visits_ok and issues_ok

    # Check if completion record already exists
    existing_completion = db.query(PlacementCompletion).filter(
        PlacementCompletion.student_id == s.id
    ).order_by(PlacementCompletion.created_at.desc()).first()

    return {
        "student_id": s.id,
        "student_name": s.full_name,
        "all_complete": all_green,
        "checklist": [
            {
                "id": "compliance",
                "label": "All 4 required compliance documents submitted",
                "ok": compliance_ok,
                "detail": compliance_detail,
            },
            {
                "id": "hours",
                "label": f"Required placement hours met ({s.required_hours:.0f}h)",
                "ok": hours_ok,
                "detail": hours_detail,
            },
            {
                "id": "visits",
                "label": f"All required workplace visits completed ({required_visits} visits)",
                "ok": visits_ok,
                "detail": visits_detail,
            },
            {
                "id": "issues",
                "label": "No open critical issues or flags",
                "ok": issues_ok,
                "detail": issues_detail,
            },
        ],
        "completion_record": {
            "id": existing_completion.id,
            "reference_number": existing_completion.reference_number,
            "completion_date": str(existing_completion.completion_date),
            "created_at": str(existing_completion.created_at),
        } if existing_completion else None,
    }


@router.post("/{student_id}/generate-completion")
def generate_placement_completion(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generates and stores a Placement Completion Record.
    Only allowed when all 4 checklist items are green.
    """
    from app.models import ComplianceDocument, Appointment, Issue, PlacementCompletion, VISIT_LIMITS
    from datetime import date

    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")

    today = date.today()
    REQUIRED_4 = [
        "working_with_children_check", "first_aid_certificate",
        "work_placement_agreement", "memorandum_of_understanding",
    ]

    submitted_types = {
        d.document_type for d in db.query(ComplianceDocument).filter(
            ComplianceDocument.student_id == s.id
        ).all()
    }
    if not all(t in submitted_types for t in REQUIRED_4):
        raise HTTPException(status_code=400, detail="Not all required compliance documents are submitted")

    if not (s.required_hours and s.completed_hours >= s.required_hours):
        raise HTTPException(status_code=400, detail="Required placement hours not yet met")

    required_visits = VISIT_LIMITS.get(s.qualification, 3)
    completed_visits = db.query(Appointment).filter(
        Appointment.student_id == s.id,
        Appointment.completed == True,
    ).count()
    if completed_visits < required_visits:
        raise HTTPException(status_code=400, detail=f"Only {completed_visits} of {required_visits} required visits completed")

    open_critical = db.query(Issue).filter(
        Issue.student_id == s.id,
        Issue.status.in_(["open", "in_progress"]),
        Issue.priority == "critical",
    ).count()
    if open_critical > 0:
        raise HTTPException(status_code=400, detail=f"Student has {open_critical} open critical issue(s)")

    # Generate sequential reference number
    count = db.query(PlacementCompletion).count() + 1
    ref = f"COMP-{today.year}-{count:05d}"

    completion = PlacementCompletion(
        student_id=s.id,
        reference_number=ref,
        completion_date=today,
        generated_by=current_user.id,
        hours_completed=s.completed_hours,
        hours_required=s.required_hours,
        compliance_docs_count=len(submitted_types),
        visits_count=completed_visits,
    )
    db.add(completion)
    db.commit()
    db.refresh(completion)

    write_audit(
        db, current_user, "placement.completion", "student",
        resource_id=s.id, resource_label=f"{s.full_name} ({s.student_id})",
        details={"reference_number": ref, "completion_date": str(today)},
    )
    db.commit()

    return {
        "id": completion.id,
        "reference_number": ref,
        "completion_date": str(today),
        "student_name": s.full_name,
        "message": f"Placement Completion Record {ref} generated successfully.",
    }


@router.delete("/{student_id}")
def delete_student(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Student not found")
    db.delete(s)
    db.commit()
    return {"message": "Student deleted"}


# ─── Issue 13 / 17 — Bulk Import from CSV/Excel ──────────────────────────────
@router.post("/bulk-import")
async def bulk_import_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Import students from a CSV or Excel (.xlsx) file.
    Expected CSV columns (header row required):
      student_id, full_name, email, phone, qualification, campus,
      status, required_hours, course_start_date, course_end_date
    Issue 13 — fixes broken Bulk Import.
    """
    filename = file.filename.lower()
    content = await file.read()

    rows = []
    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported")

    created, skipped, errors = [], [], []
    for i, row in enumerate(rows, start=2):
        sid = row.get("student_id", "").strip()
        name = row.get("full_name", "").strip()
        qual = row.get("qualification", "").strip()
        campus = row.get("campus", "sydney").strip()

        if not sid or not name or not qual:
            errors.append({"row": i, "error": "Missing required fields: student_id, full_name, qualification"})
            continue

        if qual not in QUALIFICATION_CHOICES:
            errors.append({"row": i, "error": f"Invalid qualification '{qual}'"})
            continue

        existing = db.query(Student).filter(Student.student_id == sid).first()
        if existing:
            skipped.append({"row": i, "student_id": sid, "reason": "Already exists"})
            continue

        try:
            req_hours = float(row.get("required_hours", 0) or 0)
            if req_hours == 0:
                req_hours = 288 if qual in ("CHC50121", "CHC50125") else 160

            s = Student(
                student_id=sid,
                full_name=name,
                email=row.get("email") or None,
                phone=row.get("phone") or None,
                qualification=qual,
                campus=campus,
                status=row.get("status", "active"),
                required_hours=req_hours,
                completed_hours=0,
                course_start_date=date.fromisoformat(row["course_start_date"]) if row.get("course_start_date") else None,
                course_end_date=date.fromisoformat(row["course_end_date"]) if row.get("course_end_date") else None,
            )
            db.add(s)
            created.append(sid)
        except Exception as e:
            errors.append({"row": i, "student_id": sid, "error": str(e)})

    db.commit()
    return {
        "message": f"Import complete: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors",
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
