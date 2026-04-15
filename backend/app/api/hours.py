"""
Hours Log API
Fixes:
  Issue 3  — bulk create endpoint so UI can submit multiple rows in one session
  Issue 19 — smart validation: flag shifts >10h and duplicate dates per student
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
from datetime import date, datetime

from app.database import get_db
from app.models import HoursLog, Student, User
from app.utils.auth import get_current_user

router = APIRouter()


def log_to_dict(log: HoursLog) -> dict:
    return {
        "id": log.id,
        "student_id": log.student_id,
        "log_date": str(log.log_date),
        "hours": log.hours,
        "activity_description": log.activity_description,
        "approved": log.approved,
        "approved_by": log.approved_by,
        "approved_at": str(log.approved_at) if log.approved_at else None,
        "supervisor_signed": log.supervisor_signed,
        "flagged_unrealistic": getattr(log, "flagged_unrealistic", False),
        "flagged_duplicate": getattr(log, "flagged_duplicate", False),
        "created_at": str(log.created_at) if log.created_at else None,
    }


@router.get("/summary")
def hours_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    students = db.query(Student).filter(Student.status == "active").all()
    result = []
    for s in students:
        logs = db.query(HoursLog).filter(HoursLog.student_id == s.id).all()
        approved_hours = sum(l.hours for l in logs if l.approved)
        pending_hours = sum(l.hours for l in logs if not l.approved)
        result.append({
            "student_id": s.id,
            "student_name": s.full_name,
            "student_ref": s.student_id,
            "campus": s.campus,
            "qualification": s.qualification,
            "required_hours": s.required_hours,
            "completed_hours": s.completed_hours,
            "approved_hours": approved_hours,
            "pending_hours": pending_hours,
            "percentage": round(s.completed_hours / s.required_hours * 100, 1) if s.required_hours else 0,
        })
    return result


@router.get("")
def list_hours(
    student_id: Optional[str] = None,
    approved: Optional[bool] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(HoursLog)
    if student_id:
        q = q.filter(HoursLog.student_id == student_id)
    if approved is not None:
        q = q.filter(HoursLog.approved == approved)
    if date_from:
        q = q.filter(HoursLog.log_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(HoursLog.log_date <= date.fromisoformat(date_to))
    logs = q.order_by(HoursLog.log_date.desc()).all()
    return [log_to_dict(l) for l in logs]


class HoursEntry(BaseModel):
    """A single date/hours entry (used in both single and bulk create)."""
    log_date: str
    hours: float
    activity_description: Optional[str] = None
    supervisor_signed: bool = False


class HoursCreate(BaseModel):
    student_id: str
    log_date: str
    hours: float
    activity_description: Optional[str] = None
    supervisor_signed: bool = False


def _validate_and_create(
    student: Student,
    entry_date: date,
    hours: float,
    description: Optional[str],
    supervisor_signed: bool,
    current_user_id: str,
    db: Session,
) -> HoursLog:
    """
    Validate a single log entry (Issue 19) and create the HoursLog record.
    Flags unrealistic shifts (>10h) and duplicate dates without blocking.
    """
    flagged_unrealistic = hours > 10
    # Check if there's already a log for this student on this date
    existing_today = db.query(HoursLog).filter(
        HoursLog.student_id == student.id,
        HoursLog.log_date == entry_date,
    ).first()
    flagged_duplicate = existing_today is not None

    log = HoursLog(
        student_id=student.id,
        log_date=entry_date,
        hours=hours,
        activity_description=description,
        supervisor_signed=supervisor_signed,
        approved=False,
        flagged_unrealistic=flagged_unrealistic,
        flagged_duplicate=flagged_duplicate,
        created_by=current_user_id,
    )
    return log, flagged_unrealistic, flagged_duplicate


@router.post("")
def create_hours_log(
    data: HoursCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if data.hours <= 0 or data.hours > 24:
        raise HTTPException(status_code=400, detail="Hours must be between 0 and 24")

    prev_pct = int(student.completed_hours / student.required_hours * 100) if student.required_hours else 0

    entry_date = date.fromisoformat(data.log_date)
    log, flag_unreal, flag_dup = _validate_and_create(
        student, entry_date, data.hours,
        data.activity_description, data.supervisor_signed,
        current_user.id, db,
    )
    db.add(log)
    student.completed_hours = (student.completed_hours or 0) + data.hours
    db.commit()
    db.refresh(log)

    # Milestone emails
    new_pct = int(student.completed_hours / student.required_hours * 100) if student.required_hours else 0
    for milestone in [50, 100]:
        if prev_pct < milestone <= new_pct and student.email:
            from app.services.email_service import email_hours_milestone
            email_hours_milestone(student.full_name, student.email, student.completed_hours, student.required_hours, milestone, "")

    result = log_to_dict(log)
    result["warnings"] = []
    if flag_unreal:
        result["warnings"].append("This shift exceeds 10 hours and has been flagged for review.")
    if flag_dup:
        result["warnings"].append("A log entry already exists for this student on this date.")
    return result


# ─── Issue 3 — bulk create (multiple rows in one session) ────────────────────
class HoursBulkCreate(BaseModel):
    student_id: str
    entries: List[HoursEntry]


@router.post("/bulk")
def create_hours_bulk(
    data: HoursBulkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Submit multiple placement-hour entries at once (Issue 3).
    Returns a list of created log records with any validation warnings.
    """
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if not data.entries:
        raise HTTPException(status_code=400, detail="No entries provided")

    prev_pct = int(student.completed_hours / student.required_hours * 100) if student.required_hours else 0

    results = []
    total_added = 0.0
    for entry in data.entries:
        if entry.hours <= 0 or entry.hours > 24:
            results.append({"error": f"Invalid hours ({entry.hours}) for date {entry.log_date}"})
            continue
        entry_date = date.fromisoformat(entry.log_date)
        log, flag_unreal, flag_dup = _validate_and_create(
            student, entry_date, entry.hours,
            entry.activity_description, entry.supervisor_signed,
            current_user.id, db,
        )
        db.add(log)
        total_added += entry.hours
        student.completed_hours = (student.completed_hours or 0) + entry.hours
        db.flush()
        result = log_to_dict(log)
        result["warnings"] = []
        if flag_unreal:
            result["warnings"].append(f"Shift on {entry.log_date} exceeds 10 hours — flagged for review.")
        if flag_dup:
            result["warnings"].append(f"A log entry already exists for {entry.log_date} — flagged as duplicate.")
        results.append(result)

    db.commit()

    # Milestone emails
    new_pct = int(student.completed_hours / student.required_hours * 100) if student.required_hours else 0
    for milestone in [50, 100]:
        if prev_pct < milestone <= new_pct and student.email:
            from app.services.email_service import email_hours_milestone
            email_hours_milestone(student.full_name, student.email, student.completed_hours, student.required_hours, milestone, "")

    return {"message": f"{len(results)} entries processed", "results": results}


@router.put("/{log_id}/approve")
def approve_hours(
    log_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    log = db.query(HoursLog).filter(HoursLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    log.approved = True
    log.approved_by = current_user.full_name
    log.approved_at = datetime.utcnow()
    db.commit()
    return log_to_dict(log)


@router.put("/{log_id}/reject")
def reject_hours(
    log_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    log = db.query(HoursLog).filter(HoursLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    student = db.query(Student).filter(Student.id == log.student_id).first()
    if student:
        student.completed_hours = max(0, (student.completed_hours or 0) - log.hours)
    db.delete(log)
    db.commit()
    return {"message": "Hours log rejected and removed"}


@router.delete("/{log_id}")
def delete_hours_log(
    log_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    log = db.query(HoursLog).filter(HoursLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    student = db.query(Student).filter(Student.id == log.student_id).first()
    if student:
        student.completed_hours = max(0, (student.completed_hours or 0) - log.hours)
    db.delete(log)
    db.commit()
    return {"message": "Hours log deleted"}


# ─── Issue 17 — Bulk import hours from CSV/Excel ─────────────────────────────
@router.post("/bulk-import-file")
async def bulk_import_hours_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Import hours logs from CSV or Excel.
    Expected columns: student_id (ref), log_date (YYYY-MM-DD), hours, activity_description
    """
    import csv, io as _io
    from fastapi import UploadFile, File

    content = await file.read()
    rows = []
    if file.filename.lower().endswith(".csv"):
        rows = list(csv.DictReader(_io.StringIO(content.decode("utf-8-sig"))))
    elif file.filename.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
                for row in ws.iter_rows(min_row=2, values_only=True)]
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx supported")

    created, errors = [], []
    for i, row in enumerate(rows, start=2):
        sid = row.get("student_id", "").strip()
        log_date = row.get("log_date", "").strip()
        hours_val = row.get("hours", "").strip()
        if not sid or not log_date or not hours_val:
            errors.append({"row": i, "error": "Missing student_id, log_date or hours"}); continue
        student = db.query(Student).filter(Student.student_id == sid).first()
        if not student:
            errors.append({"row": i, "error": f"Student '{sid}' not found"}); continue
        try:
            hrs = float(hours_val)
            log = HoursLog(
                student_id=student.id,
                log_date=date.fromisoformat(log_date),
                hours=hrs,
                activity_description=row.get("activity_description") or None,
                approved=False,
                flagged_unrealistic=hrs > 10,
                created_by=current_user.id,
            )
            db.add(log)
            student.completed_hours = (student.completed_hours or 0) + hrs
            created.append(f"{sid}/{log_date}")
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
    db.commit()
    return {"message": f"{len(created)} entries imported, {len(errors)} errors", "created": created, "errors": errors}
